'''
    Cashier Register App
    Author: Nguyen Pham
    Final project ECE 508
'''
from tkinter import *
from tkinter import ttk
import tkinter as tk
import tkinter.font as font
import hover_over as hover_over
import tkinter.messagebox as msg
from PIL import ImageTk,Image
import os
import smtplib
from email.message import EmailMessage
from openpyxl.reader.excel import load_workbook
from datetime import date

class Application(tk.Frame):
    def getCurrentPath(self):
        '''Function gets a working path'''
        self.path = os.path.dirname(__file__)
        self.current_path = os.chdir(self.path)
                
    def __init__(self):
        '''Function initiate the Application object'''
        self.root = tk.Tk()
        self.root.title('CASHIER APP REGISTER')
        self.root.geometry('600x600')
        self.root.resizable(width = FALSE, height = FALSE)   # set a fixed window
        self.welcomeWindow()
        self.sysEmail = 'InventorySystemNotification@gmail.com'
        self.sysPassword = 'Password1.'
        self.defaultAdminUser = 'admin'
        self.defaultAdminPwd = 'password'
        self.defaultAdminEmail = 'dvan@pdx.edu'
        self.totalSale = []
        self.data = self.parseFile('InventoryList.xlsx')  # get data info
        self.lowItem = 10
        
        
     # Parse data
    def parseFile(self, filename):
        '''Function parse data from a worksheet that store items's info'''
        self.wb = load_workbook(filename) #open the spreadsheet which is stored in the current directory
        self.sh = self.wb.active      #identify the active sheet in the file
        self.data = []      
        self.keys = ['name', 'category', 'quantity', 'price', 'ingredients'] #list of keys needed to store column in spreadsheet
        for row in range (2, (self.sh.max_row + 1)):            #offset by 1 to skip the header line
            if self.sh['B'+str(row)].value == None:      #check if item name is empty
                break                               #exit the loop if cell is empty
            else:                                   #otherwise parse data
                self.name = self.sh['B' + str(row)].value
                self.category = self.sh['C' + str(row)].value
                self.quantity = self.sh['D' + str(row)].value
                self.price = self.sh['E' + str(row)].value
                self.ingredients = self.sh['F' + str(row)].value
                self.data.append(dict(zip(self.keys, [self.name, self.category, self.quantity, self.price, self.ingredients])))
        return self.data
    
    def checkSelectItem(self):
        '''Function checks if any item selected'''
        self.key = ['name', 'price', 'quantity', 'total']
        self.selects = []
        # check which item selected
        for i in range(0,len(self.images)):        # Store selected item
            if int(self.quantity_variable[i].get()) == 0:
                continue
            else:
                self.selects.append(dict(zip(self.key,[self.name_list[i], self.price_list[i],
                                                       int(self.quantity_variable[i].get()), 0])))
        return self.selects
    
    def checkLimit(self):
        '''Function checks if selected item is over limit amount in stock'''
        limitItem= []   # store the selected item dexceed the amount in stock
        self.select_list = self.checkSelectItem()
        self.keyAtr = ['name', 'quantity']
        # check if an item selected is not more than its amount in inventory
        for select in self.select_list:
            for inventory in self.data:
                if select['name'] == inventory['name']:
                    if select['quantity'] > inventory['quantity']:   # if purchase item is more than item in the inventory
                        limitItem.append(dict(zip(self.keyAtr, [inventory['name'], inventory['quantity']])))
                    break
        return limitItem
    
    def updateClick(self):
        '''Function updates items'''
        for i in range(0, len(self.select_list)):
            self.select_list[i]['quantity'] = int(self.purchase_qty_list[i].get())
            self.select_list[i]['total'] = float(self.item_price_list[i])  
        self.purchaseWindow.destroy()
        self.payment()
     

    def paymentClick(self):
        '''Function call payment function and trace the payment() called from the menu window'''
        limit_list = self.checkLimit()
        if limit_list:#if list limitItem  is not empty ->  selected item is over amount in stock
            for element in limit_list:
                msg.showinfo( "MSG",f"Low in stock, {element['name']} has only  {element['quantity']} remaining")
        else:   
            self.call_payment = 1
            self.payment()
        
    def backWelcome_Menu(self):
        '''Function goes back to the welcomeWindow'''
        self.menuWindow.destroy()
        
    def backWelcome_Admin(self):
        '''Function goes back to the welcomeWindow'''
        self.adWindow.destroy()
        
    def backWelcome_Sale(self):
        '''Function goes back to the welcomeWindow'''
        self.reportWindow.destroy()
        self.adWindow.destroy()
    
    def backWelcome_Invoice(self):
        '''Function goes back to the welcomeWindow'''
        self.invoiceWindow.destroy()
        
    def backMenu_Payment(self):
        '''Function  goes back to the menuWindow'''
        self.purchaseWindow.destroy()
        
    def backAdminWindow_Email(self):
        '''Function goes back to the adminWidow'''
        self.emailWindow.destroy()

    def backAdminWindow_Sale(self):
        '''Function goes back to the adminWidow'''
        self.reportWindow.destroy()
    
    # #########################################
    #  #write back to the inventory on closing the app
    # def rootOnClosing(self):
    #     '''Function writes back to the inventory on closing the app'''
    #     if msg.askokcancel("Close", 'Would you like to quit?'):
    #         self.root.destroy()
    #         self.updateInventoryData(self.data)
    #     #########################################
        
    def welcomeWindow(self):
        '''Function creates a welcome page'''
        # welcome
        self.getCurrentPath()  # get the current working path
        self.welcomeImg= ImageTk.PhotoImage(Image.open('icons/welcome.png'))
        self.welcomeLabel = tk.Label(image = self.welcomeImg)
        self.welcomeLabel.place(x =10, y = 10)
        
        # get image for the start button
        self.startImg = ImageTk.PhotoImage(Image.open('icons/start.png'))
        self.imgLabel =  tk.Label(image = self.startImg)
        self.startBtt =tk.Button(self.root, image =  self.startImg, command =  self.menuWindow)
        self.startBtt.place(x= 230, y = 450)
        
        # get image for the Admin button
        self.admin_welcomeImg = ImageTk.PhotoImage(Image.open('icons/setting.png').resize((70,70), Image.ANTIALIAS))
        self.admin_welcomeLabel = tk.Label(image =   self.admin_welcomeImg)
        self.startBtt = tk.Button(self.root, image =  self.admin_welcomeImg, command =  self.adminLoginWindow)
        self.startBtt.place(x = 520, y = 520)
          
        # Current Date
        self.today = date.today()
        self.current_date = self.today.strftime("%B %d, %Y")
        self.date_label = tk.Label(self.root, text = f'{self.current_date}', font = ('Courier', 25))
        self.date_label.place(relx = 0.5, rely = 0.95, anchor = S)

         # Menu window
    def menuWindow(self):
        '''Function displays all items in the store'''
        self.call_payment = 0
        # self.data = self.parseFile('InventoryList.xlsx')  # get data info
        self.menuWindow = tk.Toplevel()
        self.menuWindow.title('MENU')
        self. menuWindow.geometry('1070x700')
        self.menuWindow.resizable(width = FALSE, height = FALSE)
        self.shape = (4,7)   # the grid of items
        self.xlen = 20
        self.ylen = 20
        self.count = 0
        self.images = []
        self.image_list=[]
        self.quantity_variable = []
        self.price_list=[]
        self.name_list = []
        self.item_select=[]
        self.ingredients=[]
        #Store the items' price
        for i in self.data:
            self.price_list.append(i['price'])
            self.name_list.append(i['name'])
            self.ingredients.append(i['ingredients'])

        # get all images from the directory
        for image in os.listdir('images'):
            if image.endswith('.png'):
                self.image_list.append(image)
        self.image_list.sort()

        # Display all the Images
        for i in range(0,len(self.image_list)):
            self.food_image = Image.open(f'images/{self.image_list[i]}')
            self.food_image = self.food_image.resize((100,100), Image.ANTIALIAS)
            self.show_image = ImageTk.PhotoImage(self.food_image)
            self.images.append(self.show_image)
            
        self.quantityOption = list(range(0,11))
        # list of variables of OptionMenu
        for i in range(0,len(self.image_list)):
            self.quantity = tk.StringVar()
            self.quantity.set('0')   # ser the default value
            self.quantity_variable.append(self.quantity)
        
        #display menu
        for i in range(0,len(self.images)):
            #image and ingredient list for each 
            self.food_label = tk.Label(self.menuWindow, image = self.images[i])
            self.food_label.place(x = self.xlen, y = self.ylen)
            hover_over.HoverWindow(self.food_label, f'Ingredients: {self.ingredients[i]}')
            
            #sale price
            self.price_label = tk.Label(self.menuWindow, text = f'Price: $ {self.price_list[i]}', font = 'Courier')
            self.price_label.place(x = self.xlen + 5, y = self.ylen + 100)
            
            #quantity drop down menu
            self.quantityLabel2 = tk.Label(self.menuWindow, text ='Quantity', font = 'Courier')
            self.quantityLabel2.place(x = self.xlen, y = self.ylen + 120)
            
            self.quantityDrop = tk.OptionMenu(self.menuWindow, self.quantity_variable[i], *self.quantityOption)
            self.quantityDrop.place(x = self.xlen + 70, y = self.ylen + 120)
            
            # position of the images
            self.count += 1
            if self.count == self.shape[1]: #if the image is the last one on the row
                self.xlen = 20
                self.ylen = self.ylen + 200
                self.count = 0
            else:
                self.xlen = self.xlen + 150
                
        # Check out button        
        self.myFont = tk.font.Font(size = 20, weight = 'bold')
        self.purchase_btt = tk.Button(self.menuWindow, text = "Go to Checkout", command = self.paymentClick,font='Helvetica', 
                              fg = 'blue',  width = 12, height = 2)
        self.purchase_btt['font'] = self.myFont
        self.purchase_btt.place(relx=0.5, rely = 0.95,  anchor=S)
        
         # Exit button
        self.exit_btt = tk.Button(self.menuWindow, text = "Exit", command = self.backWelcome_Menu,font='Helvetica', 
                              fg = 'blue',  width = 8, height = 2)
        self.exit_btt['font'] = self.myFont
        self.exit_btt.place(relx = 0, rely= 0.95, anchor=SW)
            
        self.menuWindow.mainloop()

    def payment(self):
        '''Function proceeds a payment'''
        self.purchase_qty_list =[]
        self.item_price_list = []
        if self.call_payment == 1:
            self.select_list = self.checkSelectItem() # select_list is  the selected items in menu window

        # if not, select_list is the update items
        # check if some item selected
        if len(self.select_list) != 0:
            self.price = 0
            self.total_price = 0
            
            self.purchaseWindow = tk.Toplevel()
            self.purchaseWindow.title('PAYMENT')
            self.purchaseWindow.geometry('900x800')
            self.purchaseWindow.resizable(width = FALSE, height = FALSE)
            self.quantityOption = list(range(0,11))
            
            self.purchase_label= tk.Label(self.purchaseWindow, text = "Items to Purchase", font=('Courier', 25))
            self.purchase_label.place(relx = 0.5, y = 10, anchor = N)
            
            self.line1 = tk.Label(self.purchaseWindow, text = "        ------------------------------------------------------------------------------------------------------------------------------------------")
            self.line1.place(y =40)
            
            i = 0
            self.ylen = 0
            while(i < len(self.select_list)):
                # item's name
                self.name = self.select_list[i]['name']
                self.purchase_name = tk.Label(self.purchaseWindow, text = self.name, font = ('Courier', 18))
                self.purchase_name.place(x = 50, y = self.ylen + 70)
                
                #quantiy
                self.item_quantity = self.select_list[i]['quantity']
                self.quantity_label = tk.Label(self.purchaseWindow, text = "Qty", font=('Courier', 18))
                self. quantity_label.place(x = 520, y = self.ylen + 70)
                
                self.purchase_qty = tk.StringVar(value = f'{self.item_quantity}')
                self.purchase_num = tk.OptionMenu(self.purchaseWindow, self.purchase_qty ,*self.quantityOption)
                self.purchase_num.place(x = 580, y = self.ylen + 68)
                self.purchase_qty_list.append(self.purchase_qty)  #store item quantity on the list
                
                #item price
                self.item_price = self.select_list[i]['price']                
                self.price = int(self.purchase_qty.get()) * self.item_price
                self.price = round(self.price, 2)     # round to 2 decimal digits
                self.item_price_list.append(self.price)  # store total prices for each item
                                
                self.str_out = tk.StringVar()
                self.str_out.set(f"$ {self.price}")
                self.purchase_price = tk.Label(self.purchaseWindow, textvariable = self.str_out, font = ('Courier', 20))
                self.purchase_price.place(x = 650, y = self.ylen + 65)
             
                # Update button
                self.update_btt = tk.Button(self.purchaseWindow, text = "Update", command = self.updateClick,font=('Courier', 18))
                self.update_btt.place(x = 750, y = self.ylen + 65)
            
                i += 1
                self.ylen = self.ylen + 30
                
                # Total price
                self.total_price = round((self.total_price + self.price), 2) # round to 2 decima digits
                     
            self.line2 = tk.Label(self.purchaseWindow, text = "        ------------------------------------------------------------------------------------------------------------------------------------------")
            self.line2.place(y = self.ylen + 70)
            # Total
            self.purchase_price = tk.Label(self.purchaseWindow, text = "Total", font = ('Courier', 20))
            self.purchase_price.place(x = 560, y = self.ylen + 100)
            
            self.str_out2 =tk. StringVar()
            self.str_out2.set(f"$ {self.total_price}")
            self.purchase_price = tk.Label(self.purchaseWindow, textvariable = self.str_out2, font = ('Courier', 20))
            self.purchase_price.place(x = 650, y = self.ylen + 100)
            
            # Finsh Payment button
            self.pay_btt = tk.Button(self.purchaseWindow, text = "Pay", command = self.invoiceClick,font=('Courier', 25))
            self.pay_btt.place(x = 370, y = self.ylen + 140)
            
            # Cancel button
            self.cancel_btt = tk.Button(self.purchaseWindow, text = "Cancel", command = self.backMenu_Payment,font=('Courier', 25))
            self.cancel_btt.place(x = 470, y = self.ylen + 140)
        
            self.call_payment +=  1
            
            # update total price for each item in the select_list:
            for i in range(0, len(self.select_list)):
                self.select_list[i]['total'] =  self.item_price_list[i]
            
            self.purchaseWindow.mainloop()
        # if no item selected, show an alert message
        else:
            msg.showinfo( "MSG","Please, select an item")
            self.call_payment =  1  

    def invoiceClick(self):
        '''Function prints an invoice'''
        # Do not allow to pay if updated quantity > quantity in inventory
        limitItem= []
        self.keyAtr1 = ['name', 'quantity']  
        for select in self.select_list:
            for inventory in self.data:
                if select['name'] == inventory['name']:
                    if select['quantity'] > inventory['quantity']:   # if purchase item is more than item in the inventory
                        limitItem.append(dict(zip(self.keyAtr1, [inventory['name'], inventory['quantity']])))
                    break
        if limitItem:#if list limitItem  is not empty ->  selected item is over amount in stock
            for element in  limitItem:
                msg.showinfo( "MSG",f"Low in stock, {element['name']} has only  {element['quantity']} remaining")
            self.purchaseWindow.destroy()
            self.payment()

        self.temp = []
        for i in range(0, len(self.select_list)):
            if self. select_list[i]['quantity'] == 0:
                self.temp.append(i)
            # else: 
                # self.select_list[i]['total'] = float(self.item_price_list[i])
        self.temp.reverse()
        for j in self.temp:
            del self.select_list[j]
        
        self.purchaseWindow.destroy()
        self.menuWindow.destroy()
        self.invoiceWindow= tk.Toplevel()
        self.invoiceWindow.resizable(width = FALSE, height = FALSE)
        self.invoiceWindow.title('INVOICE')
        self.invoiceWindow.geometry('600x800')
        
        # debug: Final sale item check 
        # print ('final item sale: ',self.select_list)
        # print('Transaction: ', self.total_price)
        
        i = 0
        self.ylen = 0
        
        self.line = tk.Label(self.invoiceWindow, text = " ******************************************************************************************")
        self.line.place(x = 30, y = self.ylen + 10)
        
        # Item's name
        self.invoice_label= tk.Label(self.invoiceWindow, text = "Purchased Items", font=('Courier', 16))
        self.invoice_label.place(x = 50, y = 30)
        
        # Amount
        self.invoice_label= tk.Label(self.invoiceWindow, text = "Amount", font=('Courier', 16))
        self.invoice_label.place(x = 420, y = 30)
        
        # price
        self.invoice_label= tk.Label(self.invoiceWindow, text = "Price", font=('Courier', 16))
        self.invoice_label.place(x = 510, y = 30)
        
        self.line = tk.Label(self.invoiceWindow, text = " ******************************************************************************************")
        self.line.place(x = 30, y = self.ylen + 50)
        
        while(i < len(self.select_list)):
            self.item_name = self.select_list[i]['name']
            self.purchase_name = tk.Label(self.invoiceWindow, text = f'* {self.item_name}', font = ('Courier', 14))
            self.purchase_name.place(x = 40, y = self.ylen + 80)
        
            # Quantity
            self.item_quantity = self. select_list[i]['quantity']
            # if self. select_list[i]['quantity'] < data[i]['name']
            self.purchase_num = tk.Label(self.invoiceWindow, text =  self.item_quantity,font = ('Courier', 14))
            self.purchase_num.place(x = 450, y = self.ylen + 80)
            
             #item price
            self.item_price = self.select_list[i]['total']
            self.purchase_price = tk.Label(self.invoiceWindow, text =  self.item_price,font = ('Courier', 14))
            self.purchase_price.place(x = 520, y = self.ylen + 80)
            self.ylen = self.ylen + 30

            i += 1
                
        self.line = tk.Label(self.invoiceWindow, text = "-----------------------------------")
        self.line.place(x = 380, y = self.ylen + 75)
        
         # Final price label
        self.final_Label = tk.Label(self.invoiceWindow, text = "Total", font = ('Courier', 16))
        self.final_Label.place(x = 420, y = self.ylen + 95)
        
        # Final price amount
        self.final_price = str(self.total_price)
        self.finalPrice_label = tk.Label(self.invoiceWindow, text = self.final_price,font = ('Courier', 16))
        self.finalPrice_label.place(x = 520, y = self.ylen + 95)
        
        self.line = tk.Label(self.invoiceWindow, text = "******THANK YOU FOR YOUR PURCHASE******", font = ('Courier', 16))
        self.line.place(relx = 0.5,  y = self.ylen + 155, anchor = S)

     # Cancel button
        self.done_btt = tk.Button(self.invoiceWindow, text = "Done", command = self.backWelcome_Invoice,font=('Courier', 20))
        self.done_btt.place(relx = 0.5, y = self.ylen + 195, anchor = S)        
                
        ##################################################################
        #Update totalSale list
        match = False
        temp = []       #hold items in select_list that aren't in total sale list
        lowStock = []
        if not self.totalSale:  #if total sale is empty 
            self.totalSale  = self.select_list
        else:           #if an item exists, update quantity/total sale
            for item1 in self.select_list:
                for item2 in self.totalSale:        
                    if item2['name'] == item1['name']:      
                        item2['quantity'] += item1 ['quantity']
                        item2['total'] += item2['total']
                        match = True
                        break
                    else:
                        match = False
                if not match:
                    temp.append(item1)
            if temp:    #add non-existent items to the list
                self.totalSale += temp
        
        #Update inventory data list after each transaction
        for item1 in self.select_list:
            for item2 in self.data:
                if item2['name'] == item1['name']:
                    item2['quantity'] -= item1['quantity']
                    #append item low in stock to the low stock list
                    if item2['quantity'] < self.lowItem:
                        lowStock.append(item2['name'])
                    break
       #notify admin when items are low in stock
        if lowStock:
            self.sendEmail(lowStock)
            
        #debug
        # print('low in stock: ', lowStock)
        # print ("\ntotal sale: ", self.totalSale)
        # print ("\nData list: ")
        
        # for i in self.data:
        #     print(i['name'] + '\t' +str(i['quantity']))
        
    def updateInventoryData(self, data):
        '''Write back to inventory list and to be called when the app is closed'''
        wb = load_workbook(self.inventoryDataFilename)
        sh = wb.active
        #Update only quantity in the data file
        for i in range(0,len(self.data)):
            sh['D'+ str(i+2)].value = self.data[i]['quantity'] #skip the header by adding 2
        
        wb.save(self.inventoryDataFilename)
        ##################################################################
    
     # Authorization window
    def adminLoginWindow(self):
        '''Display a login window for admin'''
        self.ad = tk.Toplevel()
        self.ad.resizable(width = FALSE, height = FALSE)
        self.ad.title('AUTHORIZATION')
        self.ad.geometry('500x400')
        self.ad.iconbitmap('icons/user_admin.ico')
        
        #get username
        self.user = tk.Label(self.ad, text = "Username", font=('Courier', 18))
        self.user.place(x = 60, y = 80)        
        self.name_inp = tk.Entry(self.ad, width = 30, borderwidth = 5)
        self.name_inp.place(x = 160, y = 80)
        
        #get password
        self.userpass = tk.Label(self.ad, text = "Password", font=('Courier', 18))
        self.userpass.place(x = 60, y = 120)
        self.pass_inp = tk.Entry(self.ad, width = 30, borderwidth = 5, show = '*')
        self.pass_inp.place(x = 160, y = 120)

        # self.login_btt = tk.Button(self.ad, text = "Login", font=('Courier', 25), command = self.loginVerification)
        self.login_btt = tk.Button(self.ad, text = "Login", font=('Courier', 25), command = self.loginVerification)
        self.login_btt.place(relx = 0.4, rely = 0.5)
        
    # Submit username and password
    def loginVerification(self):
        '''Function uses to verify admin login credentials'''
        if self.name_inp.get() == self.defaultAdminUser and self.pass_inp.get() == self.defaultAdminPwd:
            self.ad.destroy()
            self.adminWindow()  # go to adminWindow page
        else:   #notify user login failed, try again
            self.fail = tk.Label(self.ad, text='Login Fail! Please try again.', fg='red', font=('Courier', 17))
            self.fail.place(x=140, y=250)
        
    def adminWindow(self):
         '''Displays admin window after user authentication'''
         self.adWindow = tk.Toplevel()
         self.adWindow.title('ADMIN LOGISTIC')
         self.adWindow.geometry('400x500')
         self.adWindow.resizable(width = FALSE, height = FALSE)

         self.adminImg = ImageTk.PhotoImage(Image.open('icons/admin.png'))
         self.adminLabel =  tk.Label(self.adWindow, image = self.adminImg)
         self.adminLabel.place(relx = 0.5, rely = 0, anchor = N)
         
         self.emailButton = tk.Button(self.adWindow, text = 'Change Email', font=('Courier', 20), command = self.changeEmail)
         self.emailButton.place(relx = 0.29, rely = 0.65)
         self.balanceButton = tk.Button(self.adWindow, text = 'Daily Sale Report', font=('Courier', 20), command = self.saleReportWindow)
         self.balanceButton.place(relx = 0.22, rely = 0.75)
        
         self.emailExit_Button = tk.Button(self.adWindow, text='Exit', font = ('Courier', 20), width = 8, command = self.backWelcome_Admin)
         self.emailExit_Button.place(relx=0, rely=1, anchor = SW)
        
         self.adWindow.mainloop()

    def changeEmail(self):
        '''Function to allow admins update email address in the system'''
        self.emailWindow = tk.Toplevel()
        self.emailWindow.title("CHANGE ADMIN EMAIL")
        self.emailWindow.geometry("480x250")
        self.emailWindow.resizable(width = FALSE, height = FALSE)

        self.emailLabel = tk.Label(self.emailWindow, text = 'Please enter a new email address:', font = ('Courier', 15))
        self.emailLabel.place(x=50, y=40)
        
        self.emailEntry = tk.Entry(self.emailWindow, width = 40,  borderwidth = 5)
        self.emailEntry.place(x=50, y=70)
        
        self.emailButton = tk.Button(self.emailWindow, text='Enter', font = ('Courier', 20), width = 10, command = self.submitEmail)
        self.emailButton.place(x=170, y=120)

        self.emailBackButton = tk.Button(self.emailWindow, text='Back', font = ('Courier', 20), width = 8, command = self.backAdminWindow_Email)
        self.emailBackButton.place(relx=1, rely=1, anchor = SE)
        
        self.emailWindow.mainloop()
   
    def submitEmail(self):
        '''Function to save the updated email address'''
        if self.emailEntry.get() == '': # check if empty field
            self.failEmail = tk.Label(self.emailWindow, text='Please, enter your new email!', fg='red', font=('Courier', 15))
            self.failEmail.place(x=140, y=170) 
        else:
            self.defaultAdminEmail = self.emailEntry.get()
            self.emailEntry.delete(0, END)
            self.successEmail = tk.Label(self.emailWindow, text='The new email has been updated!', fg='green', font=('Courier', 15))
            self.successEmail.place(x=140, y=170)
    
    def saleReportWindow(self):
       #  '''Display a daily sale report to admin users'''    
        self.totalBalance = 0
        self.ylen1 = 40
        self.xlen1 = 30

        self.ylen2 = 30
        self.xlen2 = 30

        #####################TAB################################ 
        
        self.reportWindow = tk.Toplevel()
        self.reportWindow.title("DAILY SALE REPORT")
        self.reportWindow.geometry('800x800')
        self.reportWindow.resizable(width = FALSE, height = FALSE)
        self.tabWidget = ttk.Notebook(self.reportWindow)

        self. tabWidget.pack(pady = 5)
        
        self.tabFrame1 = tk.Frame( self.tabWidget, width= 800, height = 700)
        self.tabFrame1.pack(fill='both', expand =1)
        self.tabWidget.add(self.tabFrame1, text = "Daily Sale Report")
        
        self.tabFrame2 = tk.Frame(self.tabWidget, width= 800, height = 700)
        self.tabFrame2.pack(fill='both', expand =1)
        self.tabWidget.add(self.tabFrame2, text = "Inventory Report")

      #####################Daily Sale Report#################################
      
        self.saleLabel = tk.Label(self.tabFrame1, text = 'Daily Sale Report', font = ('Courier', 18, 'bold'))
        self.saleLabel.place(relx=0.5,rely=0, anchor = N)
        
        self.date_label2 = tk.Label(self.tabFrame1, text = f'{self.current_date}', font = ('Courier', 16))
        self.date_label2.place(relx = 0.5, y = 20, anchor = N)
        
        self.lbl1 = tk.Label(self.tabFrame1, text = "    ----------------------------------------------------------------------------------------------------------------------")
        self.lbl1.place(y = self.ylen1)

        for item in self.totalSale:
            self.nameLbl = tk.Label(self.tabFrame1, text = f'* {item["name"]}', font = ('Courier', 16))
            self.nameLbl.place(x= self.xlen1, y = self.ylen1 + 20)

            self.quantityLbl = tk.Label(self.tabFrame1, text = item['quantity'], font = ('Courier', 16))
            self.quantityLbl.place(x = self.xlen1 + 480, y=self.ylen1 + 20)

            self.totalLbl = tk.Label(self.tabFrame1, text = f"$ {item['total']}", font = ('Courier', 16))
            self.totalLbl.place(x = self.xlen1 + 550, y=self.ylen1 + 20)
            
            # Total balance
            self.totalBalance = round((self.totalBalance + item['total']), 2) # round to 2 decima digits
           
            self.ylen1 += 30
        
        self.lbl1lbl2 = tk.Label(self.tabFrame1, text = "    ---------------------------------------------------------------------------------------------------------------------")
        self.lbl1lbl2.place(y= self.ylen1 + 20)
        
        #Revenue
        self.totalBalanceLbl1 = tk.Label(self.tabFrame1, text = 'Daily Revenue: ', font = ('Courier', 16))
        self.totalBalanceLbl1.place(x = self.xlen1 + 380, y = self.ylen1 + 35)
        self.totalBalanceLbl2 = tk.Label(self.tabFrame1, text = f"$ {self.totalBalance}",font = ('Courier', 16))
        self.totalBalanceLbl2.place(x = self.xlen1 + 550, y = self.ylen1 + 35)
        ########################################################################
        
        #####################Inventory Report#################################
        self.saleLabel = tk.Label(self.tabFrame2, text = 'Inventory Report', font = ('Courier', 16, 'bold'))
        self.saleLabel.place(relx=0.5,rely=0, anchor = N)
        
        self.saleLabel = tk.Label(self.tabFrame2, text = 'Item', font = ('Courier', 16, 'bold'))
        self.saleLabel.place(x = 100, y= self.ylen2)
        
        self.saleLabel = tk.Label(self.tabFrame2, text = 'Remaining', font = ('Courier', 16, 'bold'))
        self.saleLabel.place(x = 500,y= self.ylen2)
        
        self.lbl1 = tk.Label(self.tabFrame2, text = "   ---------------------------------------------------------------------------------------------------------------------")
        self.lbl1.place(y = self.ylen2 + 20)

        for item in self.data:
            self.nameLbl = tk.Label(self.tabFrame2, text = f"* {item['name']}", font = ('Courier', 16))
            self.nameLbl.place(x= self.xlen2 + 30, y= self.ylen2 + 40)
            
            self.quantityLbl = tk.Label(self.tabFrame2, text = item['quantity'], font = ('Courier', 16))
            self.quantityLbl.place(x = self.xlen2 + 500, y=self.ylen2 + 40)
            
            self.ylen2 += 30

        ########################################################################

        self.saleBackButton = tk.Button(self.reportWindow, text='Back', font = ('Courier', 20), width = 8, command = self.backAdminWindow_Sale)
        self.saleBackButton.place(relx=1, rely=1, anchor = SE)
        
        self.saleExitButton = tk.Button(self.reportWindow, text='Exit', font = ('Courier', 20), width = 8, command = self.backWelcome_Sale)
        self.saleExitButton.place(relx=0, rely=1, anchor = SW)
        
        self.reportWindow.mainloop()

    # Send email
    def sendEmail(self, lowStock):
        #Create the email body
        self.msg = EmailMessage()
        self.msg['Subject'] = '<DO NOT REPLY> ALERT: Inventory Low In Stock'
        self.msg['From'] = self.sysEmail
        self.msg['To'] = self.defaultAdminEmail
        self.msgContent = '''\
        Hello,
    
        The following item(s) is LOW in stock. Please consider to reorder as soon as possible.
        \t%s
    
        Thank you!
    
        THIS IS A SYSTEM GENERATED MESSAGE. DO NOT REPLY
        ''' % '\n\t'.join([i.title() for i in lowStock])
    
        self.msg.set_content(self.msgContent)
    
        #Open a secure connection via SMTP to gmail
        with smtplib.SMTP_SSL('smtp.gmail.com',465) as server: #add Gmails SMTP Server
            server.ehlo()     #identify to the ESMTP server
            server.login(self.sysEmail,self.sysPassword) #login to SMTP email 
            server.send_message(self.msg) #send the message
            server.close()
            
    def start(self):
        '''Function start the app'''
        self.root.mainloop()
        
'''TEST'''
Application().start()
